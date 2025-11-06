#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成 NOUN 类型 Excel 表格
使用 named_attr_info_map 中的 NOUN 名称
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def db1_hash(hash_str):
    """实现 Rust 中的 db1_hash 函数逻辑"""
    chars = hash_str.encode('utf-8')
    if len(chars) < 1:
        return 0
    
    val = 0
    i = len(chars) - 1
    while i >= 0:
        val = (val * 27) + (chars[i] - 64)
        i -= 1
    
    # 处理溢出，模拟 Rust 的 saturating_add
    result = val + 0x81BF1
    if result > 0xFFFFFFFF:
        result = 0xFFFFFFFF
    return result & 0xFFFFFFFF

# NOUN 类型中文描述映射（常见类型）
noun_descriptions = {
    'WORL': '世界',
    'SITE': '站点',
    'ZONE': '区域',
    'EQUI': '设备',
    'PIPE': '管道',
    'BRAN': '分支',
    'STRU': '结构',
    'PANL': '面板',
    'BEAM': '梁',
    'COLU': '柱子',
    'PLAT': '平台',
    'STRA': '直管段',
    'ELBO': '弯头',
    'TEE': '三通',
    'FLAN': '法兰',
    'COMP': '组件',
    'NOZZ': '管嘴',
    'TANK': '储罐',
    'PUMP': '泵',
    'VALV': '阀门',
    'INSU': '保温',
    'FIRE': '消防',
    'CABLE': '电缆',
    'TRAY': '托盘',
    'UNDE': '未定义',
    'UDET': '用户定义类型',
    'CELL': '单元',
    'TASK': '任务',
    'SPLO': '样条',
    'USER': '用户',
    'DBSA': '数据库管理员',
    'DBSL': '数据库安全',
    'REFNO': '引用号',
    'OWNER': '所有者',
    'TYPEX': '类型扩展',
    'DIRE': '方向',
    'POSI': '位置',
    'ORIE': '方位',
    'PORS': '位置',
    'PORI': '位置',
    'GENE': '通用',
    'SCTN': '截面',
    'SPLN': '样条',
    'BEND': '弯管',
    'ACR': '弧',
    'ACRW': '弧线',
    'AEXTR': '拉伸',
    'AHU': '空气处理单元',
    'AIDGRO': '辅助组',
    'AIDLIN': '辅助线',
    'ANCI': '锚点',
    'APPLDW': '应用载荷',
    'AREADE': '面积',
    'ATLI': '属性列表',
    'ATTA': '附件',
    'ATTRRL': '属性规则',
    'BATT': '批处理',
    'BLIS': '块列表',
    'BLTA': '块表',
    'BLTP': '块属性',
    'BOX': '盒子',
    'BOXI': '盒子内部',
    'NCYL': '圆柱',
    'LPYR': '金字塔',
    'TMWL': '临时墙',
    'SREV': '旋转面',
    'PTCA': '补丁',
    'PANE': '面板',
    'SECT': '截面',
}

# 读取 JSON 文件
with open('all_attr_info.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# 从 named_attr_info_map 获取所有 NOUN 名称
noun_names = sorted(data['named_attr_info_map'].keys())

# 创建 NOUN 信息列表
noun_info = []
for noun_name in noun_names:
    # 计算对应的 NOUN ID
    noun_id = db1_hash(noun_name)
    
    # 获取中文描述
    description = noun_descriptions.get(noun_name, '未知类型')
    
    noun_info.append({
        'id': noun_id,
        'name': noun_name,
        'description': description
    })

# 创建 Excel 工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'NOUN类型表'

# 设置表头
headers = ['NOUN ID', 'NOUN 名称', '中文描述']
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 添加数据
for info in noun_info:
    ws.append([info['id'], info['name'], info['description']])

# 设置列宽
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 30

# 设置对齐方式
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for col_num, cell in enumerate(row, 1):
        if col_num == 1:  # NOUN ID 列
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

# 保存文件
output_file = 'NOUN类型表.xlsx'
wb.save(output_file)
print(f'Excel 表格已生成: {output_file}')
print(f'共包含 {len(noun_info)} 个 NOUN 类型')

# 显示前10个作为示例
print('\n前10个 NOUN 类型示例:')
for i, info in enumerate(noun_info[:10], 1):
    print(f'{i}. ID: {info["id"]}, 名称: {info["name"]}, 描述: {info["description"]}')
