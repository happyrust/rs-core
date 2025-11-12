#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成属性类型 Excel 表格
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 读取 JSON 文件
with open('all_attr_info.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# 收集所有类型
att_types = set()
default_val_types = set()

for noun_id, attrs in data['noun_attr_info_map'].items():
    for attr_id, attr in attrs.items():
        att_types.add(attr['att_type'])
        default_val_types.update(list(attr['default_val'].keys()))

# 类型中文描述映射
att_type_descriptions = {
    'BOOL': '布尔类型',
    'DIRECTION': '方向类型',
    'DOUBLE': '双精度浮点数类型',
    'ELEMENT': '元素引用类型',
    'INTEGER': '整数类型',
    'INTVEC': '整数向量类型',
    'ORIENTATION': '方向/姿态类型',
    'POSITION': '位置类型',
    'RefU64Vec': '64位无符号整数引用向量',
    'STRING': '字符串类型',
    'WORD': '单词/枚举类型',
}

default_val_type_descriptions = {
    'BoolType': '布尔类型',
    'DoubleArrayType': '双精度浮点数数组类型',
    'DoubleType': '双精度浮点数类型',
    'ElementType': '元素引用类型',
    'IntArrayType': '整数数组类型',
    'IntegerType': '整数类型',
    'RefU64Array': '64位无符号整数引用数组',
    'StringArrayType': '字符串数组类型',
    'StringType': '字符串类型',
    'Vec3Type': '三维向量类型',
    'WordType': '单词/枚举类型',
}

# 创建 Excel 工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '属性类型表'

# 设置表头
headers = ['类型名称', '中文描述', '类别']
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 添加 att_type 类型
for att_type in sorted(att_types):
    ws.append([att_type, att_type_descriptions.get(att_type, '未知类型'), 'att_type'])

# 添加 default_val 类型
for default_val_type in sorted(default_val_types):
    ws.append([default_val_type, default_val_type_descriptions.get(default_val_type, '未知类型'), 'default_val'])

# 设置列宽
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15

# 设置对齐方式
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='center')

# 保存文件
output_file = '属性类型表.xlsx'
wb.save(output_file)
print(f'Excel 表格已生成: {output_file}')
print(f'共包含 {len(att_types) + len(default_val_types)} 个类型')


























